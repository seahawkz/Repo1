from openpyxl import Workbook, load_workbook


class Book:
    def __init__(self, excel_file):
        self.book = Workbook()
        self.current_file = excel_file
        self.sheet = self.book.active
        first_columns = ['login', 'name', 'employeeId', 'choice']
        self.add_row(first_columns)

    def check_file(self, file):
        load_workbook(filename=file)

    def check_value(self, value):
        max_row = self.sheet.max_row
        for x in range(1, max_row+1):
            current_cell = self.sheet.cell(row=x, column=1)
            print(current_cell.value)
            if current_cell.value == value:
                return True

    def change_selection(self, login, value):
        max_row = self.sheet.max_row
        for x in range(1, max_row+1):
            current_cell = self.sheet.cell(row=x, column=1)
            selection_cell = self.sheet.cell(row=x, column=4)
            print(current_cell.value)
            if current_cell.value == login:
                selection_cell.value = value

    def add_row(self, data=[]):
        self.sheet.append(data)

    def get_sheet(self):
        return self.sheet

    def get_columns(self):
        return self.sheet.columns

    def get_rows(self):
        return self.sheet.rows

    def get_values(self):
        for i in self.sheet.values:
            print(i)


